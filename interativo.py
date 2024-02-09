import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl

class CriadorDePastasApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Criador de Pastas")

        # Elementos da interface
        self.label_nome_arquivo = tk.Label(root, text="Nome do Arquivo Excel:")
        self.entry_nome_arquivo = tk.Entry(root, state='readonly')
        self.button_selecionar_arquivo = tk.Button(root, text="Selecionar Arquivo", command=self.selecionar_arquivo)

        # Opção para escolher o local da coluna inteira ou da pasta
        self.label_opcao = tk.Label(root, text="Escolha o local para criar pastas:")
        self.opcao_var = tk.StringVar(value="Coluna Inteira")  # Inicia com a opção "Coluna Inteira"
        self.radio_coluna_inteira = tk.Radiobutton(root, text="Coluna Inteira", variable=self.opcao_var, value="Coluna Inteira", command=self.habilitar_opcao_coluna)
        self.radio_pasta = tk.Radiobutton(root, text="Pasta", variable=self.opcao_var, value="Pasta", command=self.executar_subprograma_pasta)

        # Botão para executar a ação escolhida
        self.button_executar = tk.Button(root, text="Executar", command=self.executar)

        # Variáveis para armazenar informações
        self.arquivo_excel = None
        self.destinos_colunas = {3: None}  # Dicionário para armazenar destinos das colunas (inicia com Coluna C)
        self.sheet = None  # Referência à planilha
        self.subprograma_pasta = None  # Referência ao subprograma

        # Layout da interface
        self.label_nome_arquivo.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
        self.entry_nome_arquivo.grid(row=0, column=1, padx=10, pady=5, columnspan=2, sticky=tk.W+tk.E)
        self.button_selecionar_arquivo.grid(row=0, column=3, padx=10, pady=5)

        self.label_opcao.grid(row=1, column=0, padx=10, pady=5, columnspan=2, sticky=tk.W)
        self.radio_coluna_inteira.grid(row=1, column=2, padx=10, pady=5, sticky=tk.W)
        self.radio_pasta.grid(row=1, column=3, padx=10, pady=5, sticky=tk.W)

        self.button_executar.grid(row=2, column=0, columnspan=4, pady=10)

    def habilitar_opcao_coluna(self):
        state_coluna = 'normal' if self.opcao_var.get() == "Coluna Inteira" else 'disabled'
        self.radio_pasta.config(state=state_coluna)

    def selecionar_arquivo(self):
        arquivo_excel = filedialog.askopenfilename(title="Selecione o Arquivo Excel", filetypes=[("Arquivos Excel", "*.xlsx;*.xls")])

        if arquivo_excel:
            self.arquivo_excel = arquivo_excel
            self.entry_nome_arquivo.config(state='normal')
            self.entry_nome_arquivo.delete(0, tk.END)
            self.entry_nome_arquivo.insert(0, arquivo_excel)
            self.entry_nome_arquivo.config(state='readonly')

    def executar(self):
        if self.opcao_var.get() == "Coluna Inteira":
            # Criar pastas da coluna inteira
            self.criar_pastas_coluna_inteira()
        elif self.opcao_var.get() == "Pasta":
            # Abrir o subprograma para a opção "Pasta"
            self.executar_subprograma_pasta()

    def criar_pastas_coluna_inteira(self):
        try:
            # Carregar o arquivo Excel
            wb = openpyxl.load_workbook(self.arquivo_excel)
            self.sheet = wb.active  # Atribuir a referência da planilha à variável

            # Iterar sobre todas as colunas
            for coluna in range(1, self.sheet.max_column + 1):
                coluna_nome = f"Coluna {coluna}"
                self.processar_coluna(coluna, coluna_nome)

            messagebox.showinfo("Conclusão", "Pastas criadas com sucesso!")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao criar pastas: {str(e)}")

    def processar_coluna(self, coluna, coluna_nome):
        try:
            # Verificar se há informações na coluna
            dados_coluna = [self.sheet.cell(row=i, column=coluna).value for i in range(1, self.sheet.max_row + 1) if self.sheet.cell(row=i, column=coluna).value is not None]

            if dados_coluna:
                # Perguntar ao usuário se deseja selecionar o destino da coluna atual
                selecionar_destino_coluna = messagebox.askyesno(f"Selecionar Destino {coluna_nome}", f"Deseja selecionar o local de destino da {coluna_nome}?")

                if selecionar_destino_coluna:
                    destino_coluna = self.perguntar_destino_coluna(coluna_nome)

                    if destino_coluna:
                        self.destinos_colunas[coluna] = destino_coluna
                        self.adicionar_coluna(coluna, pasta_destino=destino_coluna)
                        messagebox.showinfo("Sucesso", f"Pastas da {coluna_nome} criadas com sucesso!")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar coluna {coluna_nome}: {str(e)}")

    def adicionar_coluna(self, coluna, pasta_destino=None):
        if not pasta_destino:
            messagebox.showerror("Erro", "A pasta de destino não foi especificada.")
            return

        for i in range(1, self.sheet.max_row + 1):
            valor_coluna = self.sheet.cell(row=i, column=coluna).value
            nome_subpasta = str(valor_coluna).strip() if valor_coluna is not None else ""

            if nome_subpasta:
                caminho_pasta = os.path.join(pasta_destino, nome_subpasta)
                os.makedirs(caminho_pasta, exist_ok=True)

    def perguntar_destino_coluna(self, coluna_nome):
        destino_coluna = filedialog.askdirectory(title=f"Selecione a Pasta de Destino para {coluna_nome}")
        return destino_coluna if destino_coluna else None

    def executar_subprograma_pasta(self):
        # Subprograma para a opção "Pasta"
        self.subprograma_pasta = tk.Toplevel(self.root)
        self.subprograma_pasta.title("Subprograma Pasta")

        # Elementos da interface do subprograma
        label_coluna = tk.Label(self.subprograma_pasta, text="Escolha a Coluna:")
        self.coluna_combobox = ttk.Combobox(self.subprograma_pasta, values=[f"{chr(i)}" for i in range(65, 91)] + [str(i) for i in range(10)])
        label_linhas = tk.Label(self.subprograma_pasta, text="Informe as Linhas (separadas por vírgula):")
        self.entry_linhas = tk.Entry(self.subprograma_pasta)
        button_executar_pasta = tk.Button(self.subprograma_pasta, text="Executar", command=self.executar_subprograma_pasta_acao)  # Renomear aqui

        # Layout da interface do subprograma
        label_coluna.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
        self.coluna_combobox.grid(row=0, column=1, padx=10, pady=5, sticky=tk.W)
        label_linhas.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
        self.entry_linhas.grid(row=1, column=1, padx=10, pady=5, sticky=tk.W)
        button_executar_pasta.grid(row=2, column=0, columnspan=2, pady=10)

        # Adicione essa linha para garantir que a função seja chamada ao pressionar Enter
        self.subprograma_pasta.bind('<Return>', lambda event=None: button_executar_pasta.invoke())

    def executar_subprograma_pasta_acao(self):  # Renomear aqui
        coluna = self.coluna_combobox.get()
        linhas = self.entry_linhas.get()

        try:
            # Carregar o arquivo Excel
            wb = openpyxl.load_workbook(self.arquivo_excel)
            self.sheet = wb.active  # Atribuir a referência da planilha à variável

            # Perguntar ao usuário onde deseja criar as pastas
            destino_pasta = filedialog.askdirectory(title="Selecione a Pasta de Destino para as Pastas")

            if destino_pasta:
                self.adicionar_pasta(self.sheet, coluna, linhas, pasta_destino=destino_pasta)
                messagebox.showinfo("Concluído", f"Pastas criadas com sucesso!")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao criar pastas: {str(e)}")

    def adicionar_pasta(self, sheet, coluna, linhas, pasta_destino=None):
        if not pasta_destino:
            messagebox.showerror("Erro", "A pasta de destino não foi especificada.")
            return

        try:
            if coluna.isalnum() and len(coluna) == 1:
                for linha_str in linhas.split(','):
                    linha = linha_str.strip()
                    if linha.isdigit():
                        try:
                            linha_int = int(linha)
                            valor_coluna = sheet.cell(row=linha_int, column=self.get_column_index(coluna)).value
                        except (ValueError, IndexError):
                            valor_coluna = None

                        nome_subpasta = str(valor_coluna).strip() if valor_coluna is not None else ""

                        if nome_subpasta:
                            caminho_pasta = os.path.join(pasta_destino, nome_subpasta)
                            os.makedirs(caminho_pasta, exist_ok=True)

                messagebox.showinfo("Sucesso", "Pastas criadas com sucesso!")

            else:
                messagebox.showerror("Erro", "Coluna inválida. Insira uma única letra ou número.")
                return

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao criar pastas: {str(e)}")

    def get_column_index(self, coluna):
        if isinstance(coluna, str) and coluna.isalpha():
            coluna = coluna.upper()
            index = 0
            for char in coluna:
                index = index * 26 + (ord(char) - ord('A')) + 1
            return index
        elif isinstance(coluna, int) and coluna > 0:
            return coluna
        else:
            raise ValueError("Coluna inválida")

# Iniciar a aplicação
root = tk.Tk()
app = CriadorDePastasApp(root)
root.mainloop()
